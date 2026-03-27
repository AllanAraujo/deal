#!/usr/bin/env python3
"""
fill-sde-xlsx.py — Fill an SDE Calculator XLSX template with reconciled data.

Usage:
    python3 fill-sde-xlsx.py <data.json> <output.xlsx> <template.xlsx>

The JSON file contains reconciled SDE data from the /deal:sde skill.
The template is copied, filled with values and cell comments, then saved.
"""

import json
import shutil
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.comments import Comment
except ImportError:
    print("ERROR: openpyxl is not installed. Install with: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

# Row mapping: semantic key -> Excel row number
STANDARD_ROWS = {
    "sales": 3,
    "cogs": 4,
    "opex": 5,
    # Row 6 = Net Income (formula)
    "depreciation": 7,
    "interest": 8,
    "taxes": 9,
    "owner_salary": 10,
    "owner_payroll_tax": 11,
    # Row 12 = Basic SDE (formula)
}

ADDBACK_START_ROW = 15
ADDBACK_MAX_ROW = 39


def safe_string(val):
    """Prevent Excel formula injection by escaping leading trigger characters."""
    if isinstance(val, str) and val and val[0] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + val
    return val


def col_letter(year_index):
    """Year index 0->B, 1->C, 2->D."""
    return chr(ord("B") + year_index)


def fill_template(data_path, output_path, template_path):
    with open(data_path) as f:
        data = json.load(f)

    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)

    if "SDE Calculator" not in wb.sheetnames:
        print("ERROR: Template missing 'SDE Calculator' sheet", file=sys.stderr)
        sys.exit(1)

    ws = wb["SDE Calculator"]

    years = data.get("years", [])
    weights = data.get("weights", [])
    rows = data.get("rows", {})
    annotations = data.get("annotations", {})
    addbacks = rows.get("additional_addbacks", {})
    num_years = min(len(years), 3)

    # Validate array lengths
    if weights and len(weights) != num_years:
        print(f"WARNING: weights has {len(weights)} entries but {num_years} years", file=sys.stderr)

    for key in STANDARD_ROWS:
        vals = rows.get(key, [])
        if vals and len(vals) != num_years:
            print(f"WARNING: rows.{key} has {len(vals)} entries but {num_years} years", file=sys.stderr)

    # Fill year headers (B1:D1)
    for i in range(num_years):
        ws[f"{col_letter(i)}1"].value = int(years[i]) if str(years[i]).isdigit() else years[i]

    # Fill weighting percentages (B2:D2)
    for i in range(min(len(weights), num_years)):
        ws[f"{col_letter(i)}2"].value = weights[i]

    # Fill standard rows
    for key, row_num in STANDARD_ROWS.items():
        for i, val in enumerate(rows.get(key, [])[:num_years]):
            ws[f"{col_letter(i)}{row_num}"].value = val

    # Fill additional add-backs
    addback_row = ADDBACK_START_ROW
    for label, values in addbacks.items():
        if addback_row > ADDBACK_MAX_ROW:
            print(f"WARNING: Exceeded max add-back rows ({ADDBACK_MAX_ROW - ADDBACK_START_ROW + 1}). "
                  f"Skipping: {label}", file=sys.stderr)
            continue
        ws[f"A{addback_row}"].value = safe_string(label)
        for i, val in enumerate(values[:num_years]):
            ws[f"{col_letter(i)}{addback_row}"].value = val
        addback_row += 1

    # Resolve semantic annotations to cell references and add comments
    # Annotations are keyed by row name, then year string:
    #   {"sales": {"2023": {"source": "...", "status": "...", "ref": "..."}}}
    # Also support addback annotations keyed by label.
    for row_key, year_map in annotations.items():
        if not isinstance(year_map, dict):
            continue

        # Resolve row number
        if row_key in STANDARD_ROWS:
            row_num = STANDARD_ROWS[row_key]
        else:
            # Check addbacks — find the row where this label was written
            row_num = None
            check_row = ADDBACK_START_ROW
            for label in addbacks:
                if check_row > ADDBACK_MAX_ROW:
                    break
                if label == row_key:
                    row_num = check_row
                    break
                check_row += 1

        if row_num is None:
            continue

        for year_str, ann in year_map.items():
            if not isinstance(ann, dict):
                continue
            # Find column for this year
            try:
                year_idx = years.index(year_str)
            except ValueError:
                continue
            if year_idx >= 3:
                continue

            cell = ws[f"{col_letter(year_idx)}{row_num}"]
            text = "\n".join(
                f"{k.title()}: {ann[k]}" for k in ("source", "status", "note", "ref") if k in ann
            )
            if text:
                cell.comment = Comment(text, "deal:sde")

    wb.save(output_path)
    print(f"Generated: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) != 4:
        print(f"Usage: {sys.argv[0]} <data.json> <output.xlsx> <template.xlsx>", file=sys.stderr)
        sys.exit(1)

    _, data_path, output_path, template_path = sys.argv

    for label, path in [("Data file", data_path), ("Template", template_path)]:
        if not Path(path).exists():
            print(f"ERROR: {label} not found: {path}", file=sys.stderr)
            sys.exit(1)

    fill_template(data_path, output_path, template_path)
